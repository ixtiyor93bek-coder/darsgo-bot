import asyncio
import logging
import os
import io
import sqlite3
import pandas as pd
import re
from copy import copy
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string

from aiogram import Bot, Dispatcher, types, F
from aiogram.filters import CommandStart, Command, CommandObject
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import State, StatesGroup
from aiogram.fsm.storage.memory import MemoryStorage
from aiogram.types import FSInputFile, ReplyKeyboardMarkup, KeyboardButton, InlineKeyboardMarkup, InlineKeyboardButton

from aiogram.client.session.aiohttp import AiohttpSession

# =======================================================
# SOZLAMALAR
# =======================================================
# Proxy kerak bo'lsa yoqing, kerak bo'lmasa oddiy session ishlating
USE_PROXY = False

if USE_PROXY:
    bot = Bot(token="8708202570:AAHTHMINQfC-sXEmuKvLLObT8XrzGVhCaik", session=session)
else:
    bot = Bot(token="8708202570:AAHTHMINQfC-sXEmuKvLLObT8XrzGVhCaik")

dp = Dispatcher(storage=MemoryStorage())
BASE_DIR = os.path.dirname(os.path.abspath(__file__))

MASTER_ADMINS = [1652276565, 1750336134]

# =======================================================
# MA'LUMOTLAR BAZASI (SQLite)
# =======================================================
DB_PATH = os.path.join(BASE_DIR, "bot_database.db")

def init_db():
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()

    c.execute('''
        CREATE TABLE IF NOT EXISTS users (
            user_id INTEGER PRIMARY KEY,
            region TEXT,
            full_name TEXT,
            phone TEXT,
            files_processed INTEGER DEFAULT 0,
            referrer_id INTEGER,
            referrals_count INTEGER DEFAULT 0,
            is_vip INTEGER DEFAULT 0,
            joined_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')
    c.execute('''
        CREATE TABLE IF NOT EXISTS channels (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            chat_id TEXT,
            name TEXT,
            url TEXT
        )
    ''')
    c.execute('''
        CREATE TABLE IF NOT EXISTS settings (
            key TEXT PRIMARY KEY,
            value TEXT
        )
    ''')
    c.execute("INSERT OR IGNORE INTO settings (key, value) VALUES ('referral_bonus', '3')")

    c.execute('''
        CREATE TABLE IF NOT EXISTS admins (
            user_id INTEGER PRIMARY KEY,
            added_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')

    # Pending (to'plangan) import fayllar jadvali
    # Har bir foydalanuvchi uchun bir nechta fayl saqlanadi (blob sifatida)
    c.execute('''
        CREATE TABLE IF NOT EXISTS pending_files (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL,
            file_name TEXT NOT NULL,
            file_data BLOB NOT NULL,
            sinf TEXT,
            fan TEXT,
            chorak INTEGER,
            oqituvchi TEXT,
            sheet_key TEXT,
            added_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')

    conn.commit()
    conn.close()

def execute_query(query, params=(), fetch=False, fetchall=False):
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute(query, params)
    if fetch:
        res = c.fetchone()
    elif fetchall:
        res = c.fetchall()
    else:
        res = None
        conn.commit()
    conn.close()
    return res

init_db()

def get_setting(key, default):
    res = execute_query("SELECT value FROM settings WHERE key = ?", (key,), fetch=True)
    return res[0] if res else default

def is_admin(user_id):
    if user_id in MASTER_ADMINS:
        return True
    res = execute_query("SELECT 1 FROM admins WHERE user_id = ?", (user_id,), fetch=True)
    return bool(res)

# =======================================================
# PENDING FAYLLAR BOSHQARUVI
# =======================================================
def save_pending_file(user_id, file_name, file_data, sinf, fan, chorak, oqituvchi):
    """Faylni SQLite ga binary blob sifatida saqlaydi."""
    # Sheet kaliti = sinf_fan qisqartmasi (Excel sheet nomi uchun)
    sheet_key = make_sheet_name(sinf, fan)

    # Agar bu sinf+fan+chorak kombinatsiyasi allaqachon borsa, yangilash
    execute_query(
        "DELETE FROM pending_files WHERE user_id=? AND sinf=? AND fan=? AND chorak=?",
        (user_id, sinf, fan, chorak)
    )
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute(
        "INSERT INTO pending_files (user_id, file_name, file_data, sinf, fan, chorak, oqituvchi, sheet_key) VALUES (?,?,?,?,?,?,?,?)",
        (user_id, file_name, sqlite3.Binary(file_data), sinf, fan, chorak, oqituvchi, sheet_key)
    )
    conn.commit()
    conn.close()

def get_pending_files(user_id):
    """Foydalanuvchining barcha pending fayllarini qaytaradi."""
    rows = execute_query(
        "SELECT id, file_name, sinf, fan, chorak, oqituvchi, sheet_key FROM pending_files WHERE user_id=? ORDER BY sheet_key, chorak",
        (user_id,), fetchall=True
    )
    return rows or []

def get_pending_file_data(file_id):
    """Fayl binary ma'lumotini qaytaradi."""
    res = execute_query("SELECT file_data, file_name, sinf, fan, chorak, oqituvchi FROM pending_files WHERE id=?", (file_id,), fetch=True)
    return res

def clear_pending_files(user_id):
    """Foydalanuvchining barcha pending fayllarini o'chiradi (xotira tejash)."""
    execute_query("DELETE FROM pending_files WHERE user_id=?", (user_id,))

def make_sheet_name(sinf, fan):
    """Sinf va fan nomidan Excel sheet nomini hosil qiladi. Maks 31 belgi."""
    # Fan nomini qisqartirish: har so'zning 1-3 harfi
    fan_clean = re.sub(r"[:\-'\"`]", "", fan or "").strip()
    words = fan_clean.split()
    if len(words) == 1:
        abbr = words[0][:5]
    elif len(words) == 2:
        abbr = words[0][:3] + words[1][:3]
    else:
        abbr = "".join(w[:2] for w in words[:4])
    abbr = abbr[:8]

    sinf_clean = (sinf or "").replace("/", "-").strip()
    sheet_name = f"{sinf_clean}_{abbr}"

    # Excel sheet nomi: maks 31 belgi, maxsus belgilar yo'q
    for ch in r'\/*?:[]':
        sheet_name = sheet_name.replace(ch, "")
    return sheet_name[:31]

# =======================================================
# HOLATLAR (FSM)
# =======================================================
class RegState(StatesGroup):
    region = State()
    full_name = State()
    phone = State()

class AdminState(StatesGroup):
    broadcast_msg = State()
    add_channel = State()
    del_channel = State()
    set_ref_bonus = State()
    add_vip = State()
    del_vip = State()
    add_admin = State()
    del_admin = State()

# =======================================================
# TUGMALAR
# =======================================================
regions_list = [
    "Andijon", "Buxoro", "Farg'ona", "Jizzax",
    "Xorazm", "Namangan", "Navoiy", "Qashqadaryo",
    "Qoraqalpog'iston", "Samarqand", "Sirdaryo",
    "Surxondaryo", "Toshkent viloyati", "Toshkent shahri"
]

def get_regions_kb():
    kb = []
    for i in range(0, len(regions_list), 2):
        row = [KeyboardButton(text=regions_list[i])]
        if i + 1 < len(regions_list):
            row.append(KeyboardButton(text=regions_list[i+1]))
        kb.append(row)
    return ReplyKeyboardMarkup(keyboard=kb, resize_keyboard=True)

def get_phone_kb():
    return ReplyKeyboardMarkup(
        keyboard=[[KeyboardButton(text="📞 Telefon raqamni yuborish", request_contact=True)]],
        resize_keyboard=True
    )

def get_main_kb():
    return ReplyKeyboardMarkup(
        keyboard=[
            [KeyboardButton(text="📊 Mening profilim va Referal")],
            [KeyboardButton(text="📋 Yuklangan fayllar ro'yxati")],
            [KeyboardButton(text="✅ Hisobot tayyorla")],
            [KeyboardButton(text="🗑 Fayllarni tozalash")],
            [KeyboardButton(text="❓ Bot qanday ishlaydi")]
        ],
        resize_keyboard=True
    )

def get_admin_kb():
    return InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="👥 Foydalanuvchilar", callback_data="admin_users"),
         InlineKeyboardButton(text="📈 Statistika", callback_data="admin_stats")],
        [InlineKeyboardButton(text="📢 Majburiy Kanallar", callback_data="admin_channels")],
        [InlineKeyboardButton(text="🌟 VIP Boshqaruvi", callback_data="admin_vips")],
        [InlineKeyboardButton(text="👨‍💻 Adminlar", callback_data="admin_admins")],
        [InlineKeyboardButton(text="⚙️ Referal Sozlamasi", callback_data="admin_set_ref")],
        [InlineKeyboardButton(text="✉️ Xabar yuborish", callback_data="admin_broadcast")]
    ])

# =======================================================
# MAJBURIY OBUNA
# =======================================================
async def check_mandatory_subs(user_id: int):
    if is_admin(user_id): return True, None
    user = execute_query("SELECT is_vip FROM users WHERE user_id = ?", (user_id,), fetch=True)
    if user and user[0] == 1: return True, None
    channels = execute_query("SELECT chat_id, name, url FROM channels", fetchall=True)
    if not channels: return True, None
    not_subbed = []
    for ch in channels:
        try:
            member = await bot.get_chat_member(ch[0], user_id)
            if member.status in ['left', 'kicked']:
                not_subbed.append(ch)
        except Exception:
            pass
    if not_subbed:
        return False, not_subbed
    return True, None

def get_sub_kb(not_subbed):
    kb = InlineKeyboardMarkup(inline_keyboard=[])
    for ch in not_subbed:
        kb.inline_keyboard.append([InlineKeyboardButton(text=ch[1], url=ch[2])])
    kb.inline_keyboard.append([InlineKeyboardButton(text="✅ Tasdiqlash", callback_data="check_subs")])
    return kb

@dp.callback_query(F.data == "check_subs")
async def check_subs_callback(call: types.CallbackQuery):
    status, not_subbed = await check_mandatory_subs(call.from_user.id)
    if status:
        await call.message.delete()
        await call.message.answer("✅ Rahmat! Obuna tasdiqlandi.", reply_markup=get_main_kb())
    else:
        await call.answer("Iltimos, barcha kanallarga obuna bo'ling!", show_alert=True)

# =======================================================
# IMPORT FAYL PARSER
# =======================================================
def row_to_text(row):
    return " ".join([str(x) for x in row if pd.notna(x)])

def detect_subject(df):
    """Fan nomini topish — asl (original) katta-kichik harfda."""
    for i in range(min(6, len(df))):
        for j in range(len(df.columns)):
            val = df.iloc[i, j]
            if pd.isna(val) or not isinstance(val, str):
                continue
            m = re.search(r"fan\s*[:\s]\s*(.+)", val, re.IGNORECASE)
            if m:
                return m.group(1).strip()
    return "noma'lum"

def _convert_xls_to_xlsx_bytes(file_bytes):
    """XLS faylni XLSX ga aylantirib bytes qaytaradi (LibreOffice orqali)."""
    import tempfile, subprocess
    with tempfile.NamedTemporaryFile(suffix='.xls', delete=False) as tmp_in:
        tmp_in.write(file_bytes)
        tmp_in_path = tmp_in.name
    out_dir = os.path.dirname(tmp_in_path)
    try:
        subprocess.run(
            ['libreoffice', '--headless', '--convert-to', 'xlsx', '--outdir', out_dir, tmp_in_path],
            capture_output=True, timeout=30
        )
        xlsx_path = tmp_in_path.replace('.xls', '.xlsx')
        if os.path.exists(xlsx_path):
            with open(xlsx_path, 'rb') as f:
                data = f.read()
            os.remove(xlsx_path)
            return data
    finally:
        if os.path.exists(tmp_in_path):
            os.remove(tmp_in_path)
    return None


def _detect_chorak_from_row7(df):
    """
    7-qatordan (index=6) chorak raqamini aniqlaydi.
    Qoidalar:
      "1 Chorak"               -> 1
      "2 Chorak" yoki "1 yarim yillik" -> 2  (ikkalasi bo'lsa "2 Chorak" ustunligi bor)
      "3 Chorak"               -> 3
      "4 Chorak" yoki "2 yarim yillik" -> 4  (ikkalasi bo'lsa "4 Chorak" ustunligi bor)
    Qaytaradi: (chorak_num, {chorak_baho_col_index, ...})
    """
    if len(df) < 7:
        return None, {}

    row7 = df.iloc[6]  # 7-qator (0-indexed = 6)

    found = {}  # label -> col_index
    for j, val in enumerate(row7):
        if pd.isna(val):
            continue
        s = str(val).strip()
        sl = s.lower()
        # Exact / partial match
        if re.search(r'4\s*chorak', sl):
            found['4_chorak'] = j
        elif re.search(r'2\s*yarim\s*yillik', sl):
            found['2_yarim'] = j
        elif re.search(r'3\s*chorak', sl):
            found['3_chorak'] = j
        elif re.search(r'2\s*chorak', sl):
            found['2_chorak'] = j
        elif re.search(r'1\s*yarim\s*yillik', sl):
            found['1_yarim'] = j
        elif re.search(r'1\s*chorak', sl):
            found['1_chorak'] = j

    # Chorak raqamini aniqlash
    chorak_num = None
    chorak_baho_col = None

    if '4_chorak' in found:
        chorak_num = 4
        chorak_baho_col = found['4_chorak']
    elif '2_yarim' in found and '4_chorak' not in found:
        chorak_num = 4
        chorak_baho_col = found['2_yarim']

    if chorak_num is None:
        if '3_chorak' in found:
            chorak_num = 3
            chorak_baho_col = found['3_chorak']

    if chorak_num is None:
        if '2_chorak' in found:
            chorak_num = 2
            chorak_baho_col = found['2_chorak']
        elif '1_yarim' in found:
            chorak_num = 2
            chorak_baho_col = found['1_yarim']

    if chorak_num is None:
        if '1_chorak' in found:
            chorak_num = 1
            chorak_baho_col = found['1_chorak']

    return chorak_num, chorak_baho_col


def _detect_columns_from_row8(df):
    """
    8-qatordan (index=7) BSB, LI, FB, CHSB ustunlarini aniqlaydi.
    Qaytaradi: dict { bsb_cols, li_col, fb_col, chsb_cols }
    7-qatordan FB ustunini ham topib qo'shamiz.
    """
    if len(df) < 8:
        return {'bsb_cols': [], 'li_col': None, 'fb_col': None, 'chsb_cols': []}

    row8 = df.iloc[7]   # 8-qator (0-indexed=7)
    row7 = df.iloc[6]   # 7-qator (0-indexed=6) — FB uchun

    bsb_cols = []
    li_col = None
    fb_col = None
    chsb_cols = []

    for j, val in enumerate(row8):
        if pd.isna(val):
            continue
        s = str(val).strip().upper()
        if s == 'BSB':
            bsb_cols.append(j)
        elif s == 'LI':
            li_col = j
        elif s == 'CHSB':
            chsb_cols.append(j)

    # FB va Ballar (total) ni 7-qatordan qidirish
    total_col = None
    for j, val in enumerate(row7):
        if pd.isna(val):
            continue
        s = str(val).strip().upper()
        s_lower = str(val).strip().lower()
        if s == 'FB':
            fb_col = j
        elif 'ballar' in s_lower or 'jami' in s_lower:
            total_col = j

    return {
        'bsb_cols': bsb_cols,
        'li_col': li_col,
        'fb_col': fb_col,
        'chsb_cols': chsb_cols,
        'total_col': total_col
    }


def parse_import_excel(file_bytes):
    """
    Import faylidan ma'lumotlarni o'qiydi.
    Qaytaradi: dict { sinf, fan, chorak, oqituvchi, students: [{name, bsb_balls, li_ball, fb, chsb, total, chorak_baho}] }
    """
    # XLS faylni XLSX ga o'girish
    try:
        df = pd.read_excel(io.BytesIO(file_bytes), header=None, engine='openpyxl')
    except Exception:
        xlsx_bytes = _convert_xls_to_xlsx_bytes(file_bytes)
        if xlsx_bytes:
            df = pd.read_excel(io.BytesIO(xlsx_bytes), header=None, engine='openpyxl')
        else:
            df = pd.read_excel(io.BytesIO(file_bytes), header=None)

    result = {
        "sinf": None, "fan": None, "chorak": None, "oqituvchi": None,
        "students": []
    }

    # --- Sinf ---
    for i in range(min(6, len(df))):
        row_text = row_to_text(df.iloc[i])
        m = re.search(r"sinf\s*[:\s]\s*(\d{1,2})-([a-zA-Zа-яА-Я])", row_text, re.IGNORECASE)
        if m:
            result["sinf"] = f"{m.group(1)}-{m.group(2).upper()}"
            break
    if result["sinf"] is None:
        full_text = " ".join([str(x) for row in df.values for x in row if pd.notna(x)])
        m = re.search(r"\b(\d{1,2})-([a-zA-Zа-яА-Я])\b", full_text)
        if m:
            result["sinf"] = f"{m.group(1)}-{m.group(2).upper()}"

    # --- Fan nomi ---
    for i in range(min(6, len(df))):
        row_text = row_to_text(df.iloc[i])
        m = re.search(r"fan\s*[:\s]\s*(.+)", row_text, re.IGNORECASE)
        if m:
            result["fan"] = m.group(1).strip()
            break
    if result["fan"] is None:
        result["fan"] = detect_subject(df)

    # --- O'qituvchi ---
    for i in range(min(8, len(df))):
        row_text = row_to_text(df.iloc[i])
        m = re.search(r"o['`'\u2018\u2019]?qituvchi.*?fio\s*[:\s]\s*(.+)", row_text, re.IGNORECASE)
        if m:
            result["oqituvchi"] = m.group(1).strip()
            break

    # --- Chorak raqami (7-qatordan) ---
    chorak_num, chorak_baho_col = _detect_chorak_from_row7(df)
    result["chorak"] = chorak_num

    # --- 8-qatordan ustunlarni aniqlash ---
    col_info = _detect_columns_from_row8(df)
    bsb_cols = col_info['bsb_cols']      # BSB ustunlari (D, E, F uchun)
    li_col = col_info['li_col']          # LI ustuni (G uchun)
    fb_col = col_info['fb_col']          # FB ustuni (H uchun)
    chsb_cols = col_info['chsb_cols']    # CHSB ustunlari (I uchun, max qiymat)
    total_col_from_row7 = col_info.get('total_col')  # Ballar ustuni (J uchun)

    # --- O'quvchilar qatorini topish ---
    start_row = None
    for i in range(len(df)):
        val = df.iloc[i, 0]
        if pd.notna(val) and str(val).strip() == '1':
            name_val = df.iloc[i, 1] if len(df.columns) > 1 else None
            if pd.notna(name_val) and isinstance(name_val, str) and len(name_val) > 3:
                start_row = i
                break
    # Fallback: raqam sifatida 1
    if start_row is None:
        for i in range(len(df)):
            try:
                val = df.iloc[i, 0]
                if pd.notna(val) and int(float(str(val))) == 1:
                    name_val = df.iloc[i, 1] if len(df.columns) > 1 else None
                    if pd.notna(name_val) and isinstance(name_val, str) and len(name_val) > 3:
                        start_row = i
                        break
            except Exception:
                continue

    if start_row is None:
        raise Exception("O'quvchilar ro'yxati topilmadi. Jadval tuzilishini tekshiring.")

    # --- O'quvchilarni o'qish ---
    for i in range(start_row, len(df)):
        name_val = df.iloc[i, 1] if len(df.columns) > 1 else None

        if pd.isna(name_val) or not isinstance(name_val, str) or len(name_val.strip()) < 3:
            break

        student = {
            "num": df.iloc[i, 0],
            "name": name_val.strip(),
            "bsb_balls": [],   # D, E, F ustunlari
            "li_ball": None,   # G ustuni (LI)
            "fb": None,        # H ustuni (FB)
            "chsb": None,      # I ustuni (CHSB - max)
            "total": None,     # J ustuni
            "chorak_baho": None  # O ustuni
        }

        # BSB ballari — max 3 ta, D/E/F ustunlari uchun
        for col in bsb_cols[:3]:
            if col < len(df.columns):
                v = df.iloc[i, col]
                student["bsb_balls"].append(v if pd.notna(v) else None)
            else:
                student["bsb_balls"].append(None)

        # LI (Loyiha ishi) — G ustuni
        if li_col is not None and li_col < len(df.columns):
            v = df.iloc[i, li_col]
            student["li_ball"] = v if pd.notna(v) else None

        # FB — H ustuni
        if fb_col is not None and fb_col < len(df.columns):
            v = df.iloc[i, fb_col]
            student["fb"] = v if pd.notna(v) else None

        # CHSB — I ustuni (ikki CHSB bo'lsa, eng kattasi)
        chsb_values = []
        for col in chsb_cols:
            if col < len(df.columns):
                v = df.iloc[i, col]
                if pd.notna(v):
                    try:
                        chsb_values.append(float(v))
                    except Exception:
                        pass
        if chsb_values:
            student["chsb"] = max(chsb_values)

        # Jami ball — row7 dan topilgan total_col ishlatiladi
        if total_col_from_row7 is not None and total_col_from_row7 < len(df.columns):
            v = df.iloc[i, total_col_from_row7]
            student["total"] = v if pd.notna(v) else None

        # Chorak bahosi — 7-qatordan topilgan ustun
        if chorak_baho_col is not None and chorak_baho_col < len(df.columns):
            v = df.iloc[i, chorak_baho_col]
            student["chorak_baho"] = v if pd.notna(v) else None

        result["students"].append(student)

    return result

# =======================================================
# EXPORT HISOBOT GENERATSIYA
# =======================================================
BLOCK_SIZE = 49  # Har bir chorak bloki (1-49, 50-98, ...)

def safe_write(ws, row, col, value):
    """Merged cell ga yozishdan himoyalanib qiymat yozadi."""
    from openpyxl.cell.cell import MergedCell
    cell = ws.cell(row=row, column=col)
    if isinstance(cell, MergedCell):
        for rng in ws.merged_cells.ranges:
            if rng.min_row <= row <= rng.max_row and rng.min_col <= col <= rng.max_col:
                ws.cell(row=rng.min_row, column=rng.min_col).value = value
                return
        return
    cell.value = value

def chorak_row_offset(chorak_num: int) -> int:
    """N-chorak blokining boshlanish qatori (1-indexed)."""
    return (chorak_num - 1) * BLOCK_SIZE + 1

def fill_chorak_block(ws, chorak_num: int, data: dict):
    """
    Bir chorak blokini ma'lumotlar bilan to'ldiradi.
    chorak_num: 1, 2, 3, 4
    data: parse_import_excel() natijasi
    """
    base = chorak_row_offset(chorak_num)  # 1-chorak=1, 2-chorak=50, ...

    # Metadata qatorlari (base+1 = Row 2 ga mos)
    meta_row = base + 1  # Row 2

    safe_write(ws, meta_row, 2, "fan")
    safe_write(ws, meta_row, 3, data["fan"])
    safe_write(ws, meta_row, 4, "chorak")
    safe_write(ws, meta_row, 5, data["chorak"])
    safe_write(ws, meta_row, 6, "sinf ")
    safe_write(ws, meta_row, 7, data["sinf"])

    # O'qituvchi ismi (base+47 = Row 48 ga mos)
    teacher_row = base + 47
    safe_write(ws, teacher_row, 7, data["oqituvchi"])

    # O'quvchilar ma'lumotlari
    # 1-chorakda o'quvchilar Row 6 dan boshlanadi → base+5
    student_start = base + 5  # Row 6 (1-chorakda)

    for i, st in enumerate(data["students"]):
        row = student_start + i
        # C ustuni (3): ism-familiya
        safe_write(ws, row, 3, st["name"])

        # BSB ustunlari: col D, E, F (4, 5, 6)
        bsb_export_cols = [4, 5, 6]
        for j, bsb_val in enumerate(st["bsb_balls"][:3]):
            if bsb_val is not None:
                safe_write(ws, row, bsb_export_cols[j], bsb_val)

        # LI (Loyiha ishi): col G (7)
        li_val = st.get("li_ball")
        if li_val is not None:
            safe_write(ws, row, 7, li_val)

        # FB: col H (8)
        if st.get("fb") is not None:
            safe_write(ws, row, 8, st["fb"])

        # CHSB: col I (9) — max qiymat (allaqachon max sifatida saqlangan)
        if st.get("chsb") is not None:
            safe_write(ws, row, 9, st["chsb"])

        # Jami ball: col J (10)
        if st.get("total") is not None:
            safe_write(ws, row, 10, st["total"])

        # Chorak bahosi: col O (15)
        if st.get("chorak_baho") is not None:
            safe_write(ws, row, 15, st["chorak_baho"])

def generate_report(pending_rows: list) -> bytes:
    """
    Barcha pending fayllardan yagona Excel hisobot yaratadi.
    pending_rows: [(id, file_name, sinf, fan, chorak, oqituvchi, sheet_key), ...]
    """
    template_path = os.path.join(BASE_DIR, "yangi_hisobot_template.xlsx")
    if not os.path.exists(template_path):
        raise FileNotFoundError("Shablon fayl topilmadi: yangi_hisobot_template.xlsx")

    # Shablondan workbook yuklaymiz
    wb = load_workbook(template_path)

    # Mavjud sheetlarni yig'amiz (tahlil bundan mustasno)
    # Keyin kerak bo'lmagan sheetlarni o'chirib, yangisini qo'shamiz
    template_sheet_names = [s for s in wb.sheetnames if s.lower() != "tahlil"]
    tahlil_sheet = wb["tahlil"] if "tahlil" in wb.sheetnames else None

    # Shablon sifatida birinchi listni olamiz
    template_ws = wb[template_sheet_names[0]]

    # Fayllarni sinf+fan bo'yicha guruhlash
    # {sheet_key: [(chorak, data), ...]}
    groups = {}
    for row in pending_rows:
        file_id, file_name, sinf, fan, chorak, oqituvchi, sheet_key = row
        file_info = get_pending_file_data(file_id)
        if not file_info:
            continue
        file_bytes, _, _, _, _, _ = file_info
        try:
            data = parse_import_excel(bytes(file_bytes))
        except Exception as e:
            logging.warning(f"Fayl parse xatosi ({file_name}): {e}")
            continue

        if sheet_key not in groups:
            groups[sheet_key] = {"sinf": sinf, "fan": fan, "oqituvchi": oqituvchi, "quarters": []}
        groups[sheet_key]["quarters"].append((chorak, data))

    # Barcha mavjud sheetlarni o'chirib, yangisini qo'shamiz
    existing_sheets = [s for s in wb.sheetnames if s.lower() != "tahlil"]
    new_sheets = {}

    for idx, (sheet_key, group_data) in enumerate(groups.items()):
        # Birinchi sheet uchun mavjud sheetni nomlaymiz, qolganlarini ko'paytiramiz
        if idx < len(existing_sheets):
            ws = wb[existing_sheets[idx]]
            ws.title = sheet_key
        else:
            # Shablondan nusxa olamiz
            ws = wb.copy_worksheet(template_ws)
            ws.title = sheet_key

        new_sheets[sheet_key] = ws

        # Keraksiz sheetlarni o'chirish (idx > len(groups)-1 bo'lganlari)
        for old_name in existing_sheets[len(groups):]:
            if old_name in wb.sheetnames and old_name.lower() != "tahlil":
                del wb[old_name]

        # Har bir chorakni to'ldirish
        for chorak_num, chorak_data in sorted(
                group_data["quarters"],
                key=lambda x: x[0]
        ):
            try:
                fill_chorak_block(ws, chorak_num, chorak_data)
            except Exception as e:
                logging.warning(f"Chorak {chorak_num} to'ldirishda xato: {e}")

    # Tahlil sheetini yangilash
    if tahlil_sheet is not None:
        _update_tahlil_sheet(tahlil_sheet, groups, new_sheets)

    # Faylni xotiraga yozamiz
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.read()

def _update_tahlil_sheet(ws, groups: dict, sheet_map: dict):
    """
    Tahlil sheetini dinamik yangilaydi.
    Har bir sinf+fan kombinatsiyasi uchun bir qator qo'shadi.
    """
    # Sarlavha qatorlari (4, 5) ni saqlab, ma'lumot qatorlarini tozalab yozamiz
    data_start_row = 6

    # Avval mavjud ma'lumot qatorlarini tozalaymiz (6-qatordan 50-gacha)
    for r in range(data_start_row, data_start_row + 30):
        for c in range(1, 15):
            safe_write(ws, r, c, None)

    oqituvchi_name = None

    for idx, (sheet_key, group_data) in enumerate(groups.items()):
        row = data_start_row + idx
        oqituvchi_name = group_data["oqituvchi"]

        # T/r
        ws.cell(row=row, column=1).value = idx + 1

        # Sinf va fan: sheetdan formula bilan tortamiz
        # B: sinf, C: fan
        ws.cell(row=row, column=2).value = f"='{sheet_key}'!G2"
        ws.cell(row=row, column=3).value = f"='{sheet_key}'!C2"

        # Har bir chorak uchun o'rtacha ball formulasi
        # 1-chorak average: Q8 (Row 8, col Q=17)
        # 2-chorak average: Q57 (Row 57, col Q=17)
        # 3-chorak average: Q106
        # 4-chorak average: Q155
        chorak_avg_rows = {1: 8, 2: 57, 3: 106, 4: 155}
        col_map = {1: 4, 2: 5, 3: 7, 4: 9}  # D, E, G, I

        quarters_present = sorted([c for c, _ in group_data["quarters"]])

        for q_num in [1, 2, 3, 4]:
            avg_row = chorak_avg_rows[q_num]
            col = col_map[q_num]
            if q_num in quarters_present:
                ws.cell(row=row, column=col).value = f"='{sheet_key}'!Q{avg_row}"
            else:
                ws.cell(row=row, column=col).value = None

        # Farq kolonlari: F=E-D, H=G-E, J=I-G
        col_letters = {4: "D", 5: "E", 6: "F", 7: "G", 8: "H", 9: "I", 10: "J"}
        ws.cell(row=row, column=6).value = f"=IF(ISBLANK(E{row}),\"\",E{row}-D{row})"
        ws.cell(row=row, column=8).value = f"=IF(ISBLANK(G{row}),\"\",G{row}-E{row})"
        ws.cell(row=row, column=10).value = f"=IF(ISBLANK(I{row}),\"\",I{row}-G{row})"

    total_groups = len(groups)
    umumiy_row = data_start_row + total_groups + 1

    # Umumiy qator
    ws.cell(row=umumiy_row, column=1).value = "Umumiy"
    for col, q_col in [(4, "D"), (5, "E"), (7, "G"), (9, "I")]:
        r1, r2 = data_start_row, data_start_row + total_groups - 1
        ws.cell(row=umumiy_row, column=col).value = f"=IFERROR(AVERAGE({q_col}{r1}:{q_col}{r2}),\"\")"

    ws.cell(row=umumiy_row, column=6).value = f"=IF(ISBLANK(E{umumiy_row}),\"\",E{umumiy_row}-D{umumiy_row})"
    ws.cell(row=umumiy_row, column=8).value = f"=IF(ISBLANK(G{umumiy_row}),\"\",G{umumiy_row}-E{umumiy_row})"
    ws.cell(row=umumiy_row, column=10).value = f"=IF(ISBLANK(I{umumiy_row}),\"\",I{umumiy_row}-G{umumiy_row})"

    # O'qituvchi ismi va o'quv yili
    safe_write(ws, 2, 4, oqituvchi_name)
    current_year = datetime.now().year
    safe_write(ws, 2, 8, f"{current_year-1}-{current_year}")

# =======================================================
# START VA RO'YXATDAN O'TISH
# =======================================================
@dp.message(CommandStart())
async def cmd_start(message: types.Message, command: CommandObject, state: FSMContext):
    user_id = message.from_user.id
    user = execute_query("SELECT * FROM users WHERE user_id = ?", (user_id,), fetch=True)

    referrer_id = None
    if command.args and command.args.isdigit():
        referrer_id = int(command.args)
        if referrer_id == user_id: referrer_id = None

    if not user:
        await state.update_data(referrer_id=referrer_id)
        await message.answer("Assalomu alaykum! Botdan foydalanish uchun ro'yxatdan o'tishingiz kerak.\n\nIltimos, viloyatingizni tanlang:", reply_markup=get_regions_kb())
        await state.set_state(RegState.region)
    else:
        status, not_subbed = await check_mandatory_subs(user_id)
        if not status:
            await message.answer("⚠️ Botdan foydalanish uchun quyidagi kanallarga obuna bo'lishingiz shart:", reply_markup=get_sub_kb(not_subbed))
            return
        pending = get_pending_files(user_id)
        await message.answer(
            f"Assalomu alaykum yana bir bor! 👋\n\n"
            f"📁 Hozirda saqlangan fayllar: <b>{len(pending)} ta</b>\n\n"
            f"Excel faylingizni yuboring yoki hisobot tayyorlash uchun tugmani bosing.",
            reply_markup=get_main_kb(), parse_mode="HTML"
        )

@dp.message(RegState.region)
async def process_region(message: types.Message, state: FSMContext):
    if message.text not in regions_list:
        await message.answer("Iltimos, quyidagi tugmalardan birini tanlang.")
        return
    await state.update_data(region=message.text)
    await message.answer("Endi ism va familiyangizni kiriting:", reply_markup=types.ReplyKeyboardRemove())
    await state.set_state(RegState.full_name)

@dp.message(RegState.full_name)
async def process_name(message: types.Message, state: FSMContext):
    await state.update_data(full_name=message.text)
    await message.answer("Rahmat! Endi telefon raqamingizni pastdagi tugma orqali yuboring:", reply_markup=get_phone_kb())
    await state.set_state(RegState.phone)

@dp.message(RegState.phone)
async def process_phone(message: types.Message, state: FSMContext):
    if not message.contact:
        await message.answer("Iltimos, '📞 Telefon raqamni yuborish' tugmasini bosing.")
        return
    data = await state.get_data()
    user_id = message.from_user.id
    execute_query(
        "INSERT INTO users (user_id, region, full_name, phone, referrer_id) VALUES (?, ?, ?, ?, ?)",
        (user_id, data['region'], data['full_name'], message.contact.phone_number, data.get('referrer_id'))
    )
    if data.get('referrer_id'):
        execute_query("UPDATE users SET referrals_count = referrals_count + 1 WHERE user_id = ?", (data['referrer_id'],))
        try:
            await bot.send_message(data['referrer_id'], f"🎉 Sizning taklifingiz orqali {data['full_name']} botga qo'shildi!")
        except:
            pass
    await state.clear()
    status, not_subbed = await check_mandatory_subs(user_id)
    if not status:
        await message.answer("✅ Ro'yxatdan o'tdingiz!\n\nBotdan foydalanish uchun kanallarga obuna bo'ling:", reply_markup=get_sub_kb(not_subbed))
        return
    await message.answer("✅ Ro'yxatdan muvaffaqiyatli o'tdingiz!\n\nEndi Excel hisobot faylingizni yuborishingiz mumkin.", reply_markup=get_main_kb())

# =======================================================
# ASOSIY TUGMALAR
# =======================================================
@dp.message(F.text == "❓ Bot qanday ishlaydi")
async def bot_how_it_works(message: types.Message):
    text = (
        "📌 <b>Bot qanday ishlaydi:</b>\n\n"
        "1️⃣ <b>Import faylni yuboring</b> — sinf jurnali (.xls yoki .xlsx)\n"
        "2️⃣ <b>Bir nechta fayl yuborishingiz mumkin</b> — turli sinflar, fanlar va choraklar\n"
        "3️⃣ <b>'Hisobot tayyorla'</b> tugmasini bosing — barcha fayllardan yagona hisobot\n"
        "4️⃣ Hisobot tayyor bo'lgach, fayllar avtomatik tozalanadi\n\n"
        "📊 <b>Sheet nomlash:</b> sinf + fan qisqartmasi (masalan: 5-B_tarix)\n"
        "📋 <b>Tahlil:</b> oxirgi 'tahlil' listida barcha sinf/fanlar bo'yicha umumiy tahlil"
    )
    await message.answer(text, parse_mode="HTML")

@dp.message(F.text == "📊 Mening profilim va Referal")
async def show_profile(message: types.Message):
    user = execute_query("SELECT files_processed, referrals_count, is_vip FROM users WHERE user_id = ?", (message.from_user.id,), fetch=True)
    if not user:
        return
    files_processed, referrals_count, is_vip = user
    bot_info = await bot.get_me()
    ref_link = f"https://t.me/{bot_info.username}?start={message.from_user.id}"
    ref_bonus = int(get_setting('referral_bonus', 3))
    max_files = 3 + (referrals_count * ref_bonus)
    status_text = "🌟 VIP" if is_vip else "Oddiy"
    limit_text = "Cheksiz (VIP)" if is_vip else f"{max_files} ta (Hozircha: {files_processed} ta)"
    text = (
        f"👤 <b>Sizning profilingiz:</b>\n\n"
        f"💎 Status: <b>{status_text}</b>\n"
        f"📂 Ishlangan fayllar: <b>{files_processed} ta</b>\n"
        f"📈 Joriy limit: <b>{limit_text}</b>\n"
        f"👥 Taklif qilingan do'stlar: <b>{referrals_count} ta</b>\n\n"
        f"📌 <i>Har bir taklif uchun +{ref_bonus} ta fayl imkoniyati!</i>\n\n"
        f"🔗 <b>Taklif havolangiz:</b>\n{ref_link}"
    )
    await message.answer(text, parse_mode="HTML")

@dp.message(F.text == "📋 Yuklangan fayllar ro'yxati")
async def show_pending_list(message: types.Message):
    user_id = message.from_user.id
    pending = get_pending_files(user_id)
    if not pending:
        await message.answer("📭 Hozirda saqlangan fayllar yo'q.\n\nExcel fayllarini yuboring, keyin 'Hisobot tayyorla' tugmasini bosing.")
        return
    text = f"📋 <b>Saqlangan fayllar ({len(pending)} ta):</b>\n\n"
    for row in pending:
        file_id, file_name, sinf, fan, chorak, oqituvchi, sheet_key = row
        text += f"📄 <b>{sheet_key}</b> — {chorak}-chorak\n"
        text += f"   Sinf: {sinf} | Fan: {fan}\n"
        oq = oqituvchi or "noma'lum"
        text += f"   O'qituvchi: {oq}\n\n"
    text += "\n✅ Hisobot tayyorlash uchun tugmani bosing."
    await message.answer(text, parse_mode="HTML")

@dp.message(F.text == "🗑 Fayllarni tozalash")
async def clear_files_handler(message: types.Message):
    user_id = message.from_user.id
    pending = get_pending_files(user_id)
    if not pending:
        await message.answer("📭 Tozalanadigan fayl yo'q.")
        return
    kb = InlineKeyboardMarkup(inline_keyboard=[[
        InlineKeyboardButton(text="✅ Ha, o'chirish", callback_data="confirm_clear"),
        InlineKeyboardButton(text="❌ Bekor qilish", callback_data="cancel_clear")
    ]])
    await message.answer(f"⚠️ {len(pending)} ta saqlangan fayl o'chirilsinmi?", reply_markup=kb)

@dp.callback_query(F.data == "confirm_clear")
async def confirm_clear(call: types.CallbackQuery):
    clear_pending_files(call.from_user.id)
    await call.message.edit_text("✅ Barcha fayllar o'chirildi.")

@dp.callback_query(F.data == "cancel_clear")
async def cancel_clear(call: types.CallbackQuery):
    await call.message.edit_text("❌ Bekor qilindi.")

# =======================================================
# HISOBOT TAYYORLASH
# =======================================================
@dp.message(F.text == "✅ Hisobot tayyorla")
async def generate_report_handler(message: types.Message):
    user_id = message.from_user.id

    if not is_admin(user_id):
        status, not_subbed = await check_mandatory_subs(user_id)
        if not status:
            await message.answer("⚠️ Botdan foydalanish uchun kanallarga obuna bo'lishingiz shart:", reply_markup=get_sub_kb(not_subbed))
            return
        user = execute_query("SELECT files_processed, referrals_count, is_vip FROM users WHERE user_id = ?", (user_id,), fetch=True)
        if not user:
            await message.answer("Iltimos, avval /start buyrug'i orqali ro'yxatdan o'ting.")
            return
        files_processed, referrals_count, is_vip = user
        ref_bonus = int(get_setting('referral_bonus', 3))
        max_files = 3 + (referrals_count * ref_bonus)
        if not is_vip and files_processed >= max_files:
            bot_info = await bot.get_me()
            ref_link = f"https://t.me/{bot_info.username}?start={user_id}"
            await message.answer(
                f"⚠️ <b>Limitingiz tugadi!</b>\n\nDo'stlarni taklif qiling. Har bir taklif uchun <b>+{ref_bonus} ta</b> imkoniyat.\n\n🔗 {ref_link}",
                parse_mode="HTML"
            )
            return

    pending = get_pending_files(user_id)
    if not pending:
        await message.answer("📭 Hozirda saqlangan fayllar yo'q.\n\nAvval Excel fayllarini yuboring.")
        return

    msg = await message.answer(f"⏳ {len(pending)} ta fayl asosida hisobot tayyorlanmoqda...")

    try:
        output_bytes = await asyncio.to_thread(generate_report, pending)

        # Fayl nomi
        now = datetime.now().strftime("%d_%m_%Y")
        output_filename = f"Hisobot_{now}.xlsx"

        # Faylni yuborish
        doc = types.BufferedInputFile(output_bytes, filename=output_filename)
        await message.answer_document(
            document=doc,
            caption=f"✅ Hisobot tayyor!\n📊 Jami {len(pending)} ta fayl asosida tayyorlandi."
        )

        # Statistika yangilash
        execute_query("UPDATE users SET files_processed = files_processed + 1 WHERE user_id = ?", (user_id,))

        # SQLite tozalash (xotira tejash)
        clear_pending_files(user_id)
        await message.answer("🗑 Saqlangan fayllar xotiradan tozalandi.", reply_markup=get_main_kb())

    except Exception as e:
        await message.answer(f"❌ Xatolik yuz berdi:\n{str(e)}\n\nFayl tuzilishini tekshiring.")
        logging.exception(f"Hisobot generatsiya xatosi: {e}")
    finally:
        await msg.delete()

# =======================================================
# EXCEL FAYL QABUL QILISH
# =======================================================
async def loading_animation(msg: types.Message, stop_event: asyncio.Event):
    frames = ["📥 Fayl qabul qilindi, tahlil qilinmoqda.", "📥 Fayl qabul qilindi, tahlil qilinmoqda..", "📥 Fayl qabul qilindi, tahlil qilinmoqda..."]
    i = 0
    while not stop_event.is_set():
        try:
            await msg.edit_text(frames[i % len(frames)])
            i += 1
            await asyncio.sleep(0.8)
        except:
            pass

@dp.message(F.document)
async def handle_excel_document(message: types.Message):
    user_id = message.from_user.id

    if not is_admin(user_id):
        status, not_subbed = await check_mandatory_subs(user_id)
        if not status:
            await message.answer("⚠️ Botdan foydalanish uchun kanallarga obuna bo'lishingiz shart:", reply_markup=get_sub_kb(not_subbed))
            return
        user = execute_query("SELECT files_processed, referrals_count, is_vip FROM users WHERE user_id = ?", (user_id,), fetch=True)
        if not user:
            await message.answer("Iltimos, avval /start buyrug'i orqali ro'yxatdan o'ting.")
            return
        files_processed, referrals_count, is_vip = user
        ref_bonus = int(get_setting('referral_bonus', 3))
        max_files = 3 + (referrals_count * ref_bonus)
        if not is_vip and files_processed >= max_files:
            bot_info = await bot.get_me()
            ref_link = f"https://t.me/{bot_info.username}?start={user_id}"
            await message.answer(
                f"⚠️ <b>Limitingiz tugadi!</b>\n\nDo'stlarni taklif qiling. +{ref_bonus} ta imkoniyat.\n\n🔗 {ref_link}",
                parse_mode="HTML"
            )
            return

    file_name = message.document.file_name or ""
    if not (file_name.endswith('.xlsx') or file_name.endswith('.xls')):
        await message.answer("Iltimos, faqat Excel (.xlsx yoki .xls) fayl yuboring.")
        return

    msg = await message.answer("📥 Fayl qabul qilindi, tahlil qilinmoqda...")
    stop_anim = asyncio.Event()
    anim_task = asyncio.create_task(loading_animation(msg, stop_anim))

    try:
        # Faylni yuklab olish
        file_bytes_io = io.BytesIO()
        await bot.download(message.document, destination=file_bytes_io)
        file_bytes = file_bytes_io.getvalue()

        # Parse qilish
        data = await asyncio.to_thread(parse_import_excel, file_bytes)

        sinf = data.get("sinf") or "noma'lum"
        fan = data.get("fan") or "noma'lum"
        chorak = data.get("chorak")
        oqituvchi = data.get("oqituvchi") or "noma'lum"
        student_count = len(data.get("students", []))

        if chorak is None:
            stop_anim.set()
            await anim_task
            await msg.delete()
            await message.answer(
                "⚠️ <b>Chorak aniqlanmadi!</b>

"
                "Import fayl 7-qatorida '1 Chorak', '2 Chorak', '3 Chorak', '4 Chorak', "
                "'1 yarim yillik' yoki '2 yarim yillik' yozuvi topilmadi.

"
                "Fayl tuzilishini tekshiring.",
                parse_mode="HTML"
            )
            return

        # SQLite ga saqlash
        await asyncio.to_thread(save_pending_file, user_id, file_name, file_bytes, sinf, fan, chorak, oqituvchi)

        stop_anim.set()
        await anim_task
        await msg.delete()

        # Pending fayllar ro'yxati
        pending = get_pending_files(user_id)
        sheet_key = make_sheet_name(sinf, fan)

        await message.answer(
            f"✅ <b>Fayl saqlandi!</b>\n\n"
            f"📚 Fan: <b>{fan}</b>\n"
            f"🏫 Sinf: <b>{sinf}</b>\n"
            f"📅 Chorak: <b>{chorak}</b>\n"
            f"👨‍🏫 O'qituvchi: <b>{oqituvchi}</b>\n"
            f"👥 O'quvchilar: <b>{student_count} ta</b>\n"
            f"🏷 List nomi: <b>{sheet_key}</b>\n\n"
            f"📋 Jami saqlangan fayllar: <b>{len(pending)} ta</b>\n\n"
            f"Yana fayl yuboring yoki <b>'✅ Hisobot tayyorla'</b> tugmasini bosing.",
            parse_mode="HTML", reply_markup=get_main_kb()
        )

    except Exception as e:
        stop_anim.set()
        await anim_task
        await msg.delete()
        await message.answer(
            f"❌ Faylni tahlil qilishda xatolik:\n<code>{str(e)}</code>\n\n"
            f"Fayl tuzilishi to'g'riligini tekshiring.",
            parse_mode="HTML"
        )
        logging.exception(f"File parse error: {e}")

# =======================================================
# ADMIN PANEL
# =======================================================
@dp.message(Command("admin"))
async def cmd_admin(message: types.Message):
    if not is_admin(message.from_user.id): return
    await message.answer("👨‍💻 <b>Admin Panelga xush kelibsiz!</b>\n\nQuyidagi menyulardan birini tanlang:",
                         reply_markup=get_admin_kb(), parse_mode="HTML")

@dp.callback_query(F.data.startswith("admin_"))
async def admin_callbacks(call: types.CallbackQuery, state: FSMContext):
    if not is_admin(call.from_user.id): return
    action = call.data[6:]  # "admin_" dan keyingisi

    if action == "users":
        total_users = execute_query("SELECT COUNT(*) FROM users", fetch=True)[0]
        recent = execute_query("SELECT full_name, phone FROM users ORDER BY joined_at DESC LIMIT 10", fetchall=True)
        text = f"👥 <b>Barcha foydalanuvchilar: {total_users} ta</b>\n\n<b>So'nggi 10:</b>\n"
        for idx, u in enumerate(recent, 1): text += f"{idx}. {u[0]} - {u[1]}\n"
        await call.message.edit_text(text, parse_mode="HTML", reply_markup=get_admin_kb())

    elif action == "stats":
        total_users = execute_query("SELECT COUNT(*) FROM users", fetch=True)[0]
        total_files = execute_query("SELECT SUM(files_processed) FROM users", fetch=True)[0] or 0
        active_users = execute_query("SELECT COUNT(*) FROM users WHERE files_processed > 0", fetch=True)[0]
        region_stats = execute_query("SELECT region, COUNT(*) FROM users GROUP BY region ORDER BY COUNT(*) DESC LIMIT 5", fetchall=True)
        total_pending = execute_query("SELECT COUNT(*) FROM pending_files", fetch=True)[0]
        conversion = round((active_users / total_users * 100) if total_users > 0 else 0, 1)
        text = (f"📊 <b>BOT STATISTIKASI</b>\n\n"
                f"👥 Umumiy: <b>{total_users}</b>\n"
                f"📁 Jami ishlangan: <b>{total_files}</b>\n"
                f"✅ Faol foydalanuvchilar: <b>{active_users} ({conversion}%)</b>\n"
                f"⏳ Kutilayotgan fayllar: <b>{total_pending}</b>\n\n"
                f"🌍 <b>Top 5 hudud:</b>\n")
        for r in region_stats: text += f"➖ {r[0]}: {r[1]} ta\n"
        await call.message.edit_text(text, parse_mode="HTML", reply_markup=get_admin_kb())

    elif action == "channels":
        channels = execute_query("SELECT id, name, chat_id FROM channels", fetchall=True)
        text = "📢 <b>Majburiy kanallar:</b>\n\n"
        for ch in channels: text += f"ID: {ch[0]} | Nomi: {ch[1]} | ChatID: {ch[2]}\n"
        kb = InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="➕ Qo'shish", callback_data="admin_addchan"),
             InlineKeyboardButton(text="➖ O'chirish", callback_data="admin_delchan")],
            [InlineKeyboardButton(text="🔙 Orqaga", callback_data="admin_back")]
        ])
        await call.message.edit_text(text, parse_mode="HTML", reply_markup=kb)

    elif action == "vips":
        vips = execute_query("SELECT user_id, full_name FROM users WHERE is_vip=1", fetchall=True)
        text = "🌟 <b>VIP Foydalanuvchilar:</b>\n\n"
        if not vips: text += "Hech kim yo'q."
        else:
            for v in vips: text += f"ID: <code>{v[0]}</code> | {v[1]}\n"
        kb = InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="➕ VIP qo'shish", callback_data="admin_addvip"),
             InlineKeyboardButton(text="➖ VIP olish", callback_data="admin_delvip")],
            [InlineKeyboardButton(text="🔙 Orqaga", callback_data="admin_back")]
        ])
        await call.message.edit_text(text, parse_mode="HTML", reply_markup=kb)

    elif action == "admins":
        admins_db = execute_query("SELECT user_id FROM admins", fetchall=True)
        text = "👨‍💻 <b>Bot Adminlari:</b>\n\n"
        for aid in MASTER_ADMINS:
            text += f"👑 <code>{aid}</code> (Asosiy Admin)\n"
        for adm in admins_db:
            text += f"👤 <code>{adm[0]}</code> (Qo'shimcha Admin)\n"
        kb = InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="➕ Admin qo'shish", callback_data="admin_addadmin"),
             InlineKeyboardButton(text="➖ Admin o'chirish", callback_data="admin_deladmin")],
            [InlineKeyboardButton(text="🔙 Orqaga", callback_data="admin_back")]
        ])
        await call.message.edit_text(text, parse_mode="HTML", reply_markup=kb)

    elif action == "set_ref":
        current_bonus = get_setting('referral_bonus', 3)
        await call.message.edit_text(f"⚙️ Hozirgi referal bonus: <b>{current_bonus} ta fayl</b>.\n\nYangi sonni yuboring:", parse_mode="HTML")
        await state.set_state(AdminState.set_ref_bonus)

    elif action == "addchan":
        await call.message.edit_text("Format: <code>ChatID | Kanal Nomi | Ssilka</code>", parse_mode="HTML")
        await state.set_state(AdminState.add_channel)

    elif action == "delchan":
        await call.message.edit_text("O'chirmoqchi bo'lgan kanal ID sini yuboring:")
        await state.set_state(AdminState.del_channel)

    elif action == "addvip":
        await call.message.edit_text("VIP qilmoqchi bo'lgan User ID:")
        await state.set_state(AdminState.add_vip)

    elif action == "delvip":
        await call.message.edit_text("VIP olinadigan User ID:")
        await state.set_state(AdminState.del_vip)

    elif action == "addadmin":
        await call.message.edit_text("Yangi Admin User ID:")
        await state.set_state(AdminState.add_admin)

    elif action == "deladmin":
        await call.message.edit_text("O'chiriladigan Admin User ID:")
        await state.set_state(AdminState.del_admin)

    elif action == "broadcast":
        await call.message.edit_text("Xabaringizni yozing (bekor: /cancel):")
        await state.set_state(AdminState.broadcast_msg)

    elif action == "back":
        await call.message.edit_text("👨‍💻 <b>Admin Panel</b>", reply_markup=get_admin_kb(), parse_mode="HTML")

@dp.message(Command("cancel"))
async def cancel_admin_state(message: types.Message, state: FSMContext):
    if is_admin(message.from_user.id):
        await state.clear()
        await message.answer("Amal bekor qilindi.", reply_markup=get_admin_kb())

@dp.message(AdminState.add_channel)
async def process_add_channel(message: types.Message, state: FSMContext):
    try:
        chat_id, name, url = [x.strip() for x in message.text.split("|")]
        execute_query("INSERT INTO channels (chat_id, name, url) VALUES (?, ?, ?)", (chat_id, name, url))
        await message.answer(f"✅ Kanal qo'shildi: {name}", reply_markup=get_admin_kb())
        await message.answer("⚠️ Bot ushbu kanalga ADMIN qilinishi shart!")
    except Exception:
        await message.answer("❌ Noto'g'ri format! `ChatID | Nomi | Ssilka`")
    await state.clear()

@dp.message(AdminState.del_channel)
async def process_del_channel(message: types.Message, state: FSMContext):
    execute_query("DELETE FROM channels WHERE id = ?", (message.text,))
    await message.answer("✅ Kanal o'chirildi.", reply_markup=get_admin_kb())
    await state.clear()

@dp.message(AdminState.set_ref_bonus)
async def process_set_ref_bonus(message: types.Message, state: FSMContext):
    if message.text.isdigit():
        execute_query("UPDATE settings SET value = ? WHERE key = 'referral_bonus'", (message.text,))
        await message.answer(f"✅ Referal bonus: har taklif uchun {message.text} ta fayl.", reply_markup=get_admin_kb())
    else:
        await message.answer("❌ Faqat raqam yuboring.")
    await state.clear()

@dp.message(AdminState.add_vip)
async def process_add_vip(message: types.Message, state: FSMContext):
    if message.text.isdigit():
        execute_query("UPDATE users SET is_vip = 1 WHERE user_id = ?", (message.text,))
        await message.answer(f"✅ {message.text} VIP qilindi.", reply_markup=get_admin_kb())
        try: await bot.send_message(int(message.text), "🎉 Siz VIP qilib belgilandingiz!")
        except: pass
    else:
        await message.answer("❌ Faqat raqam yuboring.")
    await state.clear()

@dp.message(AdminState.del_vip)
async def process_del_vip(message: types.Message, state: FSMContext):
    if message.text.isdigit():
        execute_query("UPDATE users SET is_vip = 0 WHERE user_id = ?", (message.text,))
        await message.answer(f"✅ {message.text} VIP olib tashlandi.", reply_markup=get_admin_kb())
    else:
        await message.answer("❌ Faqat raqam yuboring.")
    await state.clear()

@dp.message(AdminState.add_admin)
async def process_add_admin(message: types.Message, state: FSMContext):
    if message.text.isdigit():
        new_id = int(message.text)
        if new_id in MASTER_ADMINS:
            await message.answer("⚠️ Bu foydalanuvchi allaqachon Asosiy Admin.")
        else:
            execute_query("INSERT OR IGNORE INTO admins (user_id) VALUES (?)", (new_id,))
            await message.answer(f"✅ {new_id} admin qilindi.", reply_markup=get_admin_kb())
            try: await bot.send_message(new_id, "🎉 Siz admin qilib tayinlandingiz! /admin")
            except: pass
    else:
        await message.answer("❌ Faqat raqam yuboring.")
    await state.clear()

@dp.message(AdminState.del_admin)
async def process_del_admin(message: types.Message, state: FSMContext):
    if message.text.isdigit():
        del_id = int(message.text)
        if del_id in MASTER_ADMINS:
            await message.answer("❌ Asosiy adminni o'chirib bo'lmaydi.", reply_markup=get_admin_kb())
        else:
            execute_query("DELETE FROM admins WHERE user_id = ?", (del_id,))
            await message.answer(f"✅ {del_id} adminlikdan olib tashlandi.", reply_markup=get_admin_kb())
    else:
        await message.answer("❌ Faqat raqam yuboring.")
    await state.clear()

@dp.message(AdminState.broadcast_msg)
async def process_broadcast(message: types.Message, state: FSMContext):
    users = execute_query("SELECT user_id FROM users", fetchall=True)
    sent_count = 0
    msg = await message.answer("⏳ Tarqatilmoqda...")
    for u in users:
        try:
            await message.copy_to(u[0])
            sent_count += 1
            await asyncio.sleep(0.05)
        except: pass
    await msg.delete()
    await state.clear()
    await message.answer(f"✅ <b>{sent_count}</b> ta foydalanuvchiga yuborildi!", parse_mode="HTML", reply_markup=get_admin_kb())

# =======================================================
# BOTNI ISHGA TUSHIRISH
# =======================================================
async def main():
    logging.basicConfig(level=logging.INFO, format="%(asctime)s - %(name)s - %(levelname)s - %(message)s")
    print("Bot muvaffaqiyatli ishga tushdi...")
    await dp.start_polling(bot)

if __name__ == "__main__":
    asyncio.run(main())
